"""AI Report Writer - Knowledge Server
FastAPI + ChromaDB + sentence-transformers + LLM
"""
import os, json, uuid, re, shutil, tempfile
from pathlib import Path
from typing import Optional
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
from dotenv import load_dotenv

# ---------- ChromaDB ----------
import chromadb
from chromadb.config import Settings

# ---------- sentence-transformers ----------
from sentence_transformers import SentenceTransformer

# ---------- Document parsing ----------
import mammoth
import openpyxl
import PyPDF2
import markdown as md_lib

# ---------- HTTP client for LLM ----------
import httpx

load_dotenv()

import os as _os
_os.environ["CHROMA_TELEMETRY_ENABLED"] = "false"
_os.environ["CHROMA_TELEMETRY_OPT_OUT"] = "true"

# ============================================================
# Config
# ============================================================
BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
CHROMA_DIR = BASE_DIR / "knowledge" / "chroma_db"
FRONTEND_DIR = BASE_DIR / "frontend"

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
CHROMA_DIR.mkdir(parents=True, exist_ok=True)

LLM_API_KEY = os.getenv("LLM_API_KEY", "")
LLM_API_URL = os.getenv("LLM_API_URL", "https://api.openai.com/v1")
LLM_MODEL = os.getenv("LLM_MODEL", "gpt-4o-mini")
HOST = os.getenv("HOST", "0.0.0.0")
PORT = int(os.getenv("PORT", "8888"))

# ============================================================
# Init
# ============================================================
app = FastAPI(title="AI Report Writer")

# ChromaDB
chroma_client = chromadb.PersistentClient(path=str(CHROMA_DIR))
collection = chroma_client.get_or_create_collection(name="documents")

# Embedding model (lazy load - only used if preload failed)
_embedder = _embedder if "_embedder" in dir() else None

def get_embedder():
    global _embedder
    if _embedder is None:
        print("[INFO] Loading embedding model...")
        _embedder = SentenceTransformer("all-MiniLM-L6-v2")
    return _embedder

# Session store (in-memory)
sessions = {}

# ============================================================
# Document Parsing
# ============================================================
def parse_docx(path):
    with open(path, "rb") as f:
        result = mammoth.extract_raw_text(f)
    return result.value or ""

def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    texts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else "" for c in row]
            rows.append(vals)
            texts.append(" | ".join(vals))
        # Store structured data
        sheet_key = f"table_{sheet}"
    wb.close()
    return "\n".join(texts)

def parse_pdf(path):
    texts = []
    with open(path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
    return "\n".join(texts)

def parse_txt(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def parse_file(file_path: Path) -> dict:
    ext = file_path.suffix.lower()
    if ext == ".docx":
        text = parse_docx(file_path)
    elif ext == ".xlsx":
        text = parse_xlsx(file_path)
    elif ext == ".pdf":
        text = parse_pdf(file_path)
    else:
        text = parse_txt(file_path)
    
    return {"text": text, "fileType": ext}

# ============================================================
# Chunking
# ============================================================
def chunk_text(text: str, max_chars: int = 500) -> list:
    paragraphs = re.split(r"\n\s*\n", text)
    chunks = []
    current = ""
    for p in paragraphs:
        p = p.strip()
        if not p:
            continue
        if len(current) + len(p) < max_chars:
            current += ("\n" + p) if current else p
        else:
            if current:
                chunks.append(current)
            current = p
    if current:
        chunks.append(current)
    return chunks if chunks else [text]

# ============================================================
# LLM API
# ============================================================
async def call_llm(prompt: str, system: str = "", temperature: float = 0.1) -> str:
    if not LLM_API_KEY:
        return json.dumps({"error": "LLM API Key 未配置"})
    
    messages = []
    if system:
        messages.append({"role": "system", "content": system})
    messages.append({"role": "user", "content": prompt})
    
    async with httpx.AsyncClient(timeout=60) as client:
        try:
            resp = await client.post(
                f"{LLM_API_URL}/chat/completions",
                headers={
                    "Authorization": f"Bearer {LLM_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": LLM_MODEL,
                    "messages": messages,
                    "temperature": temperature,
                }
            )
            resp.raise_for_status()
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            return json.dumps({"error": str(e)})

async def analyze_document(text: str, file_name: str) -> dict:
    system = """你是一个文档分析助手。分析文档内容，返回 JSON（不要 markdown 包裹）：
{
  "summary": "200字以内摘要",
  "outline": ["章节1", "章节2", ...],
  "viewpoints": [
    {"title": "观点标题", "evidence": "支撑证据", "confidence": 0.xx}
  ],
  "tables": [{"name": "表名", "headers": ["列1"], "rows": [["值"]]}]
}"""
    prompt = f"文档名称：{file_name}\n\n文档内容（前 8000 字）：\n{text[:8000]}"
    result = await call_llm(prompt, system, 0.3)
    try:
        # Try to extract JSON from response
        json_match = re.search(r'\{.*\}', result, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
    except:
        pass
    return {"summary": text[:200], "outline": [], "viewpoints": [], "tables": []}

async def detect_intent(text: str, knowledge_summary: str) -> str:
    system = """判断用户意图，只返回以下一个词：
- 数据分析（查数据、看图表、做统计）
- 核心观点提炼（提炼观点、找关键信号）
- 报告撰写（写报告、生成报告）
- 自由对话（以上都不是）"""
    prompt = f"知识库摘要：{knowledge_summary}\n\n用户问题：{text}"
    return (await call_llm(prompt, system, 0.1)).strip()

async def rag_query(question: str, chunks: list) -> str:
    system = "基于知识库内容回答问题。如果知识库没有相关信息，说'当前文档中没有相关信息'。引用格式：[文件名]"
    context = "\n---\n".join([f"[{c['source']}]\n{c['text']}" for c in chunks])
    prompt = f"知识库：\n{context}\n\n问题：{question}"
    return await call_llm(prompt, system, 0.2)

# ============================================================
# API Routes
# ============================================================
@app.get("/", response_class=HTMLResponse)
async def root():
    html_path = FRONTEND_DIR / "index.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>前端文件缺失</h1><p>请确保 frontend/index.html 存在</p>")

# --- LLM Config ---
@app.post("/api/llm/config")
async def update_llm_config(data: dict):
    global LLM_API_KEY, LLM_API_URL, LLM_MODEL
    LLM_API_KEY = data.get("api_key", LLM_API_KEY)
    LLM_API_URL = data.get("api_url", LLM_API_URL)
    LLM_MODEL = data.get("model", LLM_MODEL)
    return {"status": "ok", "message": "LLM 配置已更新"}

@app.get("/api/llm/status")
async def llm_status():
    return {
        "configured": bool(LLM_API_KEY),
        "api_url": LLM_API_URL,
        "model": LLM_MODEL
    }

# --- Knowledge / Upload ---
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), folder: str = Form("default")):
    if not file.filename:
        raise HTTPException(400, "文件名无效")
    
    ext = Path(file.filename).suffix.lower()
    if ext not in [".docx", ".xlsx", ".pdf", ".txt", ".md"]:
        raise HTTPException(400, f"不支持的文件格式: {ext}")
    
    # Save
    file_id = str(uuid.uuid4())[:8]
    save_name = f"{file_id}_{file.filename}"
    save_path = UPLOAD_DIR / save_name
    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Parse
    parsed = parse_file(save_path)
    text = parsed["text"]
    if not text.strip():
        raise HTTPException(400, "文件内容为空或无法解析")
    
    # Chunk
    chunks = chunk_text(text)
    
    # Vectorize & store
    embedder = get_embedder()
    embeddings = embedder.encode(chunks).tolist()
    
    ids = [f"{file_id}_{i}" for i in range(len(chunks))]
    metadatas = [{
        "source": file.filename,
        "chunk": i,
        "text": chunks[i][:100]
    } for i in range(len(chunks))]
    
    # Remove old docs with same source name
    existing = collection.get(where={"source": file.filename})
    if existing and existing["ids"]:
        collection.delete(ids=existing["ids"])
    
    collection.add(
        documents=chunks,
        embeddings=embeddings,
        metadatas=metadatas,
        ids=ids
    )
    
    # LLM analyze
    analysis = await analyze_document(text, file.filename)
    
    # Track file in folder
    try:
        fdata = load_folders()
        target_folder = folder
        # If folder is "default", use first non-demo folder or demo folder
        if target_folder == "default":
            for f in fdata["folders"]:
                if not f.get("is_demo"):
                    target_folder = f["id"]
                    break
            if target_folder == "default":
                target_folder = "demo"
        
        for f in fdata["folders"]:
            if f["id"] == target_folder:
                if file.filename not in f.get("files", []):
                    f.setdefault("files", []).append(file.filename)
                break
        save_folders(fdata)
    except Exception as e:
        print(f"[WARN] Failed to update folder: {e}")
    
    return {
        "status": "ok",
        "file": {
            "id": file_id,
            "name": file.filename,
            "size": len(content),
            "chunks": len(chunks)
        },
        "analysis": analysis
    }

@app.get("/api/knowledge/list")
async def knowledge_list():
    all_data = collection.get()
    files = {}
    for meta in (all_data["metadatas"] or []):
        src = meta.get("source", "unknown")
        if src not in files:
            files[src] = {"name": src, "chunks": 0}
        files[src]["chunks"] += 1
    return {"files": list(files.values())}

@app.delete("/api/knowledge/{file_name:path}")
async def delete_knowledge(file_name: str):
    existing = collection.get(where={"source": file_name})
    if existing and existing["ids"]:
        collection.delete(ids=existing["ids"])
    return {"status": "ok", "deleted": len(existing.get("ids", []))}

@app.get("/api/knowledge/search")
async def search_knowledge(q: str = Query("", description="搜索关键词"), top_k: int = 5):
    if not q:
        return {"results": []}
    embedder = get_embedder()
    q_vec = embedder.encode([q]).tolist()
    results = collection.query(query_embeddings=q_vec, n_results=top_k)
    items = []
    if results and results["metadatas"]:
        for i, meta in enumerate(results["metadatas"][0]):
            items.append({
                "source": meta.get("source", ""),
                "text": results["documents"][0][i][:300] if results["documents"] else "",
                "score": results["distances"][0][i] if results.get("distances") else 0
            })
    return {"results": items}


# --- Folder Management ---
FOLDERS_FILE = BASE_DIR / "knowledge" / "folders.json"

def load_folders():
    if FOLDERS_FILE.exists():
        with open(FOLDERS_FILE, "r") as f:
            return json.load(f)
    return {"folders": [{"id": "demo", "name": "演示", "files": [], "is_demo": True}]}

def save_folders(folders_data):
    FOLDERS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(FOLDERS_FILE, "w") as f:
        json.dump(folders_data, f, ensure_ascii=False, indent=2)

@app.get("/api/folders/list")
async def list_folders():
    data = load_folders()
    # Enrich each folder with file info from chromadb
    all_data = collection.get()
    meta_map = {}
    for meta in (all_data["metadatas"] or []):
        src = meta.get("source", "unknown")
        meta_map.setdefault(src, 0)
        meta_map[src] += 1
    
    for folder in data["folders"]:
        enriched_files = []
        for fname in folder.get("files", []):
            enriched_files.append({
                "name": fname,
                "chunks": meta_map.get(fname, 0)
            })
        folder["file_details"] = enriched_files
    
    return {"status": "ok", "data": data}

@app.post("/api/folders/create")
async def create_folder(data: dict):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(400, "文件夹名称不能为空")
    
    folders_data = load_folders()
    # Check duplicate
    for f in folders_data["folders"]:
        if f["name"] == name:
            raise HTTPException(400, f"文件夹 '{name}' 已存在")
    
    new_id = "folder_" + str(uuid.uuid4())[:8]
    folders_data["folders"].append({
        "id": new_id,
        "name": name,
        "files": [],
        "is_demo": False
    })
    save_folders(folders_data)
    return {"status": "ok", "folder": {"id": new_id, "name": name}}

@app.delete("/api/folders/{folder_id}")
async def delete_folder(folder_id: str):
    folders_data = load_folders()
    folders_data["folders"] = [f for f in folders_data["folders"] if f["id"] != folder_id]
    save_folders(folders_data)
    return {"status": "ok"}

@app.post("/api/demo/seed")
async def seed_demo_data():
    """Seed demo knowledge base with example data"""
    try:
        demo_texts = {
            "北京NEV市场分析.docx": "北京新能源市场调研数据：\n北京人口2183.2万，30-44岁占比29.1%。\n26-40岁占70%以上，本科及以上学历过半，硕士及以上15.5%。\n白领占比47.1%，高收入群体40.6%，中高收入38.5%。\n人均可支配收入8.5万元。\n核心活动区为朝阳区、海淀区。\n居住区集中在丰台、通州等近郊区。\n高频商圈：朝阳大悦城（首选），高频景区：奥林匹克公园、朝阳公园、圆明园。\n线下业态以公园与购物中心为主，职住模式为核心城区活动+近郊区居住。\nAPP偏好：微信、腾讯视频、高德地图渗透率超75%。\n主力车型为中级轿车，Top3为秦L（性价比）、Model 3（科技感）、小米SU7（全域智能）。\n66%用户月均驾驶46次以上（高频驾驶）。\n需求关键词：长续航、快充效率、冬季低温适应性。\n最满意空间体验，改进点包括数字座舱卡顿、音响噪音、智能辅助驾驶稳定性。\n消费升级趋势：从基础实用向科技赋能和场景化品质进阶，聚焦通勤+家庭两大场景。\n企业启示：需强化长续航、快充、大空间、智驾、座舱，精准续航标定，搭建车主社群（科技沙龙+亲子露营），强化北京区域售后服务。",
            "中国汽车出口月报.txt": "2025年11月中国汽车出口数据：\n当月汽车出口量：28.4万辆，同比增长12.3%。\n1-11月累计出口：312.5万辆，同比增长18.7%。\n新能源汽车出口占比：38.2%，其中纯电占比62%，插混占比38%。\n主要出口目的地：俄罗斯（18.2%）、墨西哥（9.5%）、泰国（7.8%）、巴西（6.3%）、比利时（5.1%）。\n出口品牌Top5：上汽（21.3%）、比亚迪（15.7%）、奇瑞（13.2%）、吉利（11.8%）、长城（8.9%）。\n新能源出口均价：2.86万美元，同比提升8.3%。\n智能驾驶出口法规风险：欧盟关税壁垒、美国IRA法案影响、东南亚右舵市场改造需求。\n出口模式转型：从整车出口向KD散件组装+本地化生产转型，海外产能布局加速。\n2025年预计全年出口超340万辆，连续两年全球第一。"
        }
        
        embedder = get_embedder()
        
        for fname, text in demo_texts.items():
            existing = collection.get(where={"source": fname})
            if existing and existing.get("ids"):
                continue
            chunks = chunk_text(text)
            file_id = "demo_" + str(uuid.uuid4())[:8]
            embeddings = embedder.encode(chunks).tolist()
            ids = [f"{file_id}_{i}" for i in range(len(chunks))]
            metadatas = [{"source": fname, "chunk": i, "text": chunks[i][:100]} for i in range(len(chunks))]
            collection.add(documents=chunks, embeddings=embeddings, metadatas=metadatas, ids=ids)
        
        folders_data = load_folders()
        for folder in folders_data["folders"]:
            if folder["id"] == "demo":
                for fname in demo_texts:
                    if fname not in folder.get("files", []):
                        folder.setdefault("files", []).append(fname)
                break
        save_folders(folders_data)
        
        return {"status": "ok", "message": f"已导入 {len(demo_texts)} 个示例文件", "files": list(demo_texts.keys())}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"status": "error", "message": str(e)}

@app.post("/api/folders/scan")
async def scan_local_folder(data: dict):
    local_path = data.get("path", "").strip()
    if not local_path or not os.path.isdir(local_path):
        raise HTTPException(400, f"无效的文件夹路径: {local_path}")
    
    supported_exts = {".docx", ".xlsx", ".pdf", ".txt", ".md"}
    found_files = []
    
    for fname in os.listdir(local_path):
        ext = Path(fname).suffix.lower()
        if ext in supported_exts:
            found_files.append(fname)
    
    return {
        "status": "ok",
        "data": {
            "path": local_path,
            "files": found_files,
            "count": len(found_files)
        }
    }

@app.post("/api/folders/import")
async def import_folder_files(data: dict):
    local_path = data.get("path", "").strip()
    folder_id = data.get("folder_id", "")
    file_names = data.get("files", [])
    
    if not local_path or not os.path.isdir(local_path):
        raise HTTPException(400, "无效的文件夹路径")
    
    embedder = get_embedder()
    imported = []
    
    for fname in file_names:
        fpath = Path(local_path) / fname
        if not fpath.exists():
            continue
        
        try:
            parsed = parse_file(fpath)
            text = parsed["text"]
            if not text.strip():
                continue
            
            chunks = chunk_text(text)
            file_id = "import_" + str(uuid.uuid4())[:8]
            embeddings = embedder.encode(chunks).tolist()
            ids = [f"{file_id}_{i}" for i in range(len(chunks))]
            metadatas = [{
                "source": fname,
                "chunk": i,
                "text": chunks[i][:100]
            } for i in range(len(chunks))]
            
            existing = collection.get(where={"source": fname})
            if existing and existing["ids"]:
                collection.delete(ids=existing["ids"])
            
            collection.add(documents=chunks, embeddings=embeddings, metadatas=metadatas, ids=ids)
            imported.append(fname)
        except Exception as e:
            print(f"[WARN] Failed to import {fname}: {e}")
    
    # Update folder
    if folder_id:
        folders_data = load_folders()
        for folder in folders_data["folders"]:
            if folder["id"] == folder_id:
                for fname in imported:
                    if fname not in folder.get("files", []):
                        folder.setdefault("files", []).append(fname)
                break
        save_folders(folders_data)
    
    return {"status": "ok", "imported": imported, "count": len(imported)}
@app.post("/api/chat")
async def chat(data: dict):
    text = data.get("text", "").strip()
    session_id = data.get("session_id", "default")
    
    if not text:
        return {"status": "error", "message": "请输入内容"}
    
    # Init session
    if session_id not in sessions:
        sessions[session_id] = {
            "history": [],
            "adopted_viewpoints": [],
            "viewpoint_candidates": []
        }
    session = sessions[session_id]
    
    # Knowledge summary
    all_docs = collection.get()
    knowledge_summary = f"知识库共 {len(all_docs.get('ids', []))} 个片段"
    
    # Detect intent
    intent = await detect_intent(text, knowledge_summary)
    
    # RAG: retrieve relevant chunks
    embedder = get_embedder()
    q_vec = embedder.encode([text]).tolist()
    rag_results = collection.query(query_embeddings=q_vec, n_results=5)
    retrieved = []
    if rag_results and rag_results["metadatas"] and rag_results["metadatas"][0]:
        for i, meta in enumerate(rag_results["metadatas"][0]):
            retrieved.append({
                "source": meta.get("source", ""),
                "text": rag_results["documents"][0][i] if rag_results["documents"] else ""
            })
    
    # Build response based on intent
    if intent == "数据分析":
        # Return data tables from document analysis
        tables = []
        charts = []
        if retrieved:
            tables.append({
                "name": "相关数据",
                "source": retrieved[0]["source"] if retrieved else "",
                "headers": ["数据项", "值"],
                "rows": [["来自", retrieved[0]["source"]]]
            })
        
        return {
            "status": "ok",
            "data": {
                "mode": "数据分析",
                "intent": {"original_input": text},
                "data": {
                    "tables": tables,
                    "charts": charts,
                    "datasets": [{"name": "知识库"}]
                },
                "response": f"📊 已查询知识库。找到 {len(retrieved)} 个相关片段。"
            }
        }
    
    elif intent == "核心观点提炼":
        # Get pre-analyzed viewpoints from session or generate
        if not session["viewpoint_candidates"]:
            # Try to generate from retrieved content
            context = "\n".join([r["text"][:500] for r in retrieved[:3]])
            sys_prompt = "根据以下内容提炼 3-5 条核心观点。返回 JSON 数组：[{\"title\":\"...\", \"evidence\":\"...\", \"confidence\":0.xx}]"
            vp_result = await call_llm(f"内容：{context}", sys_prompt, 0.3)
            try:
                json_match = re.search(r'\[.*\]', vp_result, re.DOTALL)
                if json_match:
                    candidates = json.loads(json_match.group())
                    for i, c in enumerate(candidates):
                        c["id"] = f"vp_{i}"
                        c["source"] = retrieved[0]["source"] if retrieved else "知识库"
                    session["viewpoint_candidates"] = candidates
            except:
                pass
        
        if not session["viewpoint_candidates"]:
            session["viewpoint_candidates"] = [
                {"id": "vp_0", "title": "基于文档自动分析", "evidence": "请先上传文档", "confidence": 0.5, "source": "系统"}
            ]
        
        return {
            "status": "ok",
            "data": {
                "mode": "核心观点提炼",
                "insight": {
                    "candidates": session["viewpoint_candidates"],
                    "total": len(session["viewpoint_candidates"])
                }
            }
        }
    
    elif intent == "报告撰写":
        return {
            "status": "ok",
            "data": {
                "mode": "报告撰写",
                "report": {
                    "report_type": "分析报告",
                    "title": "分析报告",
                    "chapters": [
                        {"title": "引言", "desc": "背景与目的", "content": "请上传文档或采纳观点后生成内容。"},
                        {"title": "分析", "desc": "核心分析", "content": "请上传文档或采纳观点后生成内容。"},
                        {"title": "结论", "desc": "总结与建议", "content": "请上传文档或采纳观点后生成内容。"}
                    ]
                }
            }
        }
    
    else:
        # 自由对话：RAG
        answer = text
        if retrieved:
            answer = await rag_query(text, retrieved[:3])
        else:
            answer = "知识库为空，请先上传文档。"
        
        return {
            "status": "ok",
            "data": {
                "mode": "自由对话",
                "intent": {"original_input": text},
                "response": answer
            }
        }

@app.post("/api/insight/filter")
async def filter_insight(data: dict):
    session_id = data.get("session_id", "default")
    signal_id = data.get("signal_id", "")
    action = data.get("action", "")
    
    session = sessions.get(session_id)
    if not session:
        return {"status": "error", "message": "会话不存在"}
    
    for vp in session["viewpoint_candidates"]:
        if vp.get("id") == signal_id:
            if action == "accept":
                vp["adopted"] = True
                if vp not in session["adopted_viewpoints"]:
                    session["adopted_viewpoints"].append(vp)
            break
    
    return {"status": "ok"}

@app.post("/api/insight/card")
async def insight_card(data: dict):
    session_id = data.get("session_id", "default")
    session = sessions.get(session_id)
    if not session:
        return {"status": "error", "data": {"cards": [], "count": 0}}
    
    cards = [{
        "title": vp.get("title", ""),
        "evidence": vp.get("evidence", ""),
        "confidence": vp.get("confidence", 0.5),
        "source": vp.get("source", "")
    } for vp in session.get("adopted_viewpoints", [])]
    
    return {"status": "ok", "data": {"cards": cards, "count": len(cards)}}

@app.post("/api/report/generate")
async def report_generate(data: dict):
    session_id = data.get("session_id", "default")
    session = sessions.get(session_id)
    
    adopted = session.get("adopted_viewpoints", []) if session else []
    chapters = []
    
    for i, vp in enumerate(adopted):
        chapters.append({
            "title": vp.get("title", f"第{i+1}章"),
            "desc": vp.get("evidence", "")[:50],
            "content": f"基于观点：{vp.get('title', '')}\n\n{vp.get('evidence', '')}\n\n（内容由 AI 基于知识库生成，请双击编辑完善。）"
        })
    
    if not chapters:
        chapters = [
            {"title": "概述", "desc": "项目背景", "content": "请采纳观点后重新生成。"},
            {"title": "分析", "desc": "核心分析", "content": "请采纳观点后重新生成。"},
        ]
    
    return {
        "status": "ok",
        "data": {
            "title": "分析报告",
            "chapters": chapters
        }
    }

# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    import uvicorn
    # Pre-load model
    print("[INFO] Pre-loading embedding model...")
    try:
        m = get_embedder()
        print(f"[INFO] Model loaded: {m.get_sentence_embedding_dimension()}d")
    except Exception as e:
        print(f"[WARN] Failed to pre-load model: {e}")
    
    print(f"""
╔══════════════════════════════════════╗
║     AI Report Writer Server         ║
║     http://localhost:{PORT}          ║
╚══════════════════════════════════════╝
    """)
    uvicorn.run(app, host=HOST, port=PORT, log_level="info")
